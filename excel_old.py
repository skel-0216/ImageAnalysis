import openpyxl
from openpyxl import Workbook

import asset_filename
from fish2hist import *

# src_path_ = 'C:/Users/kms27/Desktop/GitHub/ImageAnalysis/source/'
src_path_ = './source/'
experiment_ = 'croaker#1'
"""
stage_ = 'stage00/'
size_ = '250/'
position_ = 'stomach/'
brightness_ = 'dark/'
"""
img_num = ['00.png', '01.png', '02.png', '03.png', '04.png']

sheet_name = experiment_[:-1]


def write_mackerel(path_, work_sheet):
    work_sheet.append([path_])
    for i in range(1, 6):
        temp_path = path_ + img_num[i]
        temp_list = sum(mackerel2hist_H(cv2.imread(temp_path)).tolist(), [])

        result_list = [img_num[i][:2] + ' mackerel'] + temp_list

        work_sheet.append(result_list)
    work_sheet.append(['', ''])
    return


def write_squid(title_, path_, work_sheet):
    work_sheet.append([title_])
    cnt = 0
    for i in range(0, 5):
        bgr = ['b', 'g', 'r']
        for paths in path_:
            for path in paths:
                cnt += 1
                print(path)
                for j in range(3):
                    temp_list = sum(squid2hist_BGR(cv2.imread(path))[j].tolist(), [])

                    result_list = [img_num[i][:2] + ' croaker ' + bgr[j]] + temp_list

                    work_sheet.append(result_list)
                work_sheet.append(['', ''])
        work_sheet.append(['', ''])
    work_sheet.append(['', ''])
    work_sheet.append(['', ''])
    return


def write_mackerel_stop(title_, path_, work_sheet):
    work_sheet.append([title_])
    print(path_[0])
    for i in range(len(path_[0])):
        stage = 0
        for temp_path in path_:
            print(temp_path[i])
            temp_list = sum(mackerel2hist_H(cv2.imread(temp_path[i])).tolist(), [])

            result_list = ['mackerel_%d_stage%d' % (i, stage)] + temp_list
            stage += 1

            work_sheet.append(result_list)
        work_sheet.append(['', ''])
    work_sheet.append(['', ''])
    return


ws = 0
wb = Workbook()

ws1_1 = wb.active
title_1_1 = 'croaker_body_bright'
ws1_1.title = title_1_1
write_squid(title_1_1, asset_filename.croaker1_body_bright, ws1_1)

title_1_2 = 'mackerel#4_body_dark'
ws1_2 = wb.create_sheet(title_1_2)
write_squid(title_1_2, asset_filename.croaker1_body_dark, ws1_2)

wb.save("./excel/croaker_test.xlsx")


"""
ws = 0
wb = Workbook()

ws1_1 = wb.active
title_1_1 = 'mackerel#4_continue_back_bright'
ws1_1.title = title_1_1
write_mackerel_stop(title_1_1, asset_filename.mackerel4_continue_back_bright, ws1_1)

title_1_2 = 'mackerel#4_continue_back_dark'
ws1_2 = wb.create_sheet(title_1_2)
write_mackerel_stop(title_1_2, asset_filename.mackerel4_continue_back_dark, ws1_2)

title_2_1 = 'mackerel#4_continue_belly_bright'
ws2_1 = wb.create_sheet(title_2_1)
write_mackerel_stop(title_2_1, asset_filename.mackerel4_continue_belly_dark, ws2_1)

title_2_2 = 'mackerel#4_continue_belly_dark'
ws2_2 = wb.create_sheet(title_2_2)
write_mackerel_stop(title_2_2, asset_filename.mackerel4_continue_belly_dark, ws2_2)


wb.save("./excel/mackerel#4_continue_body.xlsx")
"""


"""
ws = 0
wb = Workbook()

ws1_1 = wb.active
title_1_1 = 'mackerel#4_stop_back_bright'
ws1_1.title = title_1_1
write_mackerel_stop(title_1_1, asset_filename.mackerel4_stop_back_bright, ws1_1)

title_1_2 = 'mackerel#4_stop_back_dark'
ws1_2 = wb.create_sheet(title_1_2)
write_mackerel_stop(title_1_2, asset_filename.mackerel4_stop_back_dark, ws1_2)

title_2_1 = 'mackerel#4_stop_belly_bright'
ws2_1 = wb.create_sheet(title_2_1)
write_mackerel_stop(title_2_1, asset_filename.mackerel4_stop_belly_dark, ws2_1)

title_2_2 = 'mackerel#4_stop_belly_dark'
ws2_2 = wb.create_sheet(title_2_2)
write_mackerel_stop(title_2_2, asset_filename.mackerel4_stop_belly_dark, ws2_2)


wb.save("./excel/mackerel#4_stop.xlsx")
"""

"""
# 엑셀파일 쓰기
wb = Workbook()

ws1_1 = wb.active

ws1_1.title = 'mackerel#2_dark'

write_mackerel('source/mackerel#2/stage00/250/stomach/dark/', ws1_1)
write_mackerel('source/mackerel#2/stage01/250/stomach/dark/', ws1_1)
write_mackerel('source/mackerel#2/stage00/350/stomach/dark/', ws1_1)
write_mackerel('source/mackerel#2/stage01/350/stomach/dark/', ws1_1)

write_mackerel('source/mackerel#2/stage00/250/back/dark/', ws1_1)
write_mackerel('source/mackerel#2/stage01/250/back/dark/', ws1_1)
write_mackerel('source/mackerel#2/stage00/350/back/dark/', ws1_1)
write_mackerel('source/mackerel#2/stage01/350/back/dark/', ws1_1)

ws1_2 = wb.create_sheet('mackerel#2_bright')

write_mackerel('source/mackerel#2/stage00/250/stomach/bright/', ws1_2)
write_mackerel('source/mackerel#2/stage01/250/stomach/bright/', ws1_2)
write_mackerel('source/mackerel#2/stage00/350/stomach/bright/', ws1_2)
write_mackerel('source/mackerel#2/stage01/350/stomach/bright/', ws1_2)

write_mackerel('source/mackerel#2/stage00/250/back/bright/', ws1_2)
write_mackerel('source/mackerel#2/stage01/250/back/bright/', ws1_2)
write_mackerel('source/mackerel#2/stage00/350/back/bright/', ws1_2)
write_mackerel('source/mackerel#2/stage01/350/back/bright/', ws1_2)


ws2_1 = wb.create_sheet('mackerel#3_dark')

write_mackerel('source/mackerel#3/stage00/250/stomach/dark/', ws2_1)
write_mackerel('source/mackerel#3/stage01/250/stomach/dark/', ws2_1)
write_mackerel('source/mackerel#3/stage00/350/stomach/dark/', ws2_1)
write_mackerel('source/mackerel#3/stage01/350/stomach/dark/', ws2_1)

write_mackerel('source/mackerel#3/stage00/250/back/dark/', ws2_1)
write_mackerel('source/mackerel#3/stage01/250/back/dark/', ws2_1)
write_mackerel('source/mackerel#3/stage00/350/back/dark/', ws2_1)
write_mackerel('source/mackerel#3/stage01/350/back/dark/', ws2_1)

ws2_2 = wb.create_sheet('mackerel#3_bright')

write_mackerel('source/mackerel#3/stage00/250/stomach/bright/', ws2_2)
write_mackerel('source/mackerel#3/stage01/250/stomach/bright/', ws2_2)
write_mackerel('source/mackerel#3/stage00/350/stomach/bright/', ws2_2)
write_mackerel('source/mackerel#3/stage01/350/stomach/bright/', ws2_2)

write_mackerel('source/mackerel#3/stage00/250/back/bright/', ws2_2)
write_mackerel('source/mackerel#3/stage01/250/back/bright/', ws2_2)
write_mackerel('source/mackerel#3/stage00/350/back/bright/', ws2_2)
write_mackerel('source/mackerel#3/stage01/350/back/bright/', ws2_2)

ws3_1 = wb.create_sheet('squid#1_dark')

write_squid('source/squid#1/stage00/head/dark/', ws3_1)
write_squid('source/squid#1/stage01/head/dark/', ws3_1)
write_squid('source/squid#1/stage02/head/dark/', ws3_1)

write_squid('source/squid#1/stage00/body/dark/', ws3_1)
write_squid('source/squid#1/stage01/body/dark/', ws3_1)
write_squid('source/squid#1/stage02/body/dark/', ws3_1)


ws3_2 = wb.create_sheet('squid#1_bright')

write_squid('source/squid#1/stage00/body/bright/', ws3_2)
write_squid('source/squid#1/stage01/body/bright/', ws3_2)
write_squid('source/squid#1/stage02/body/bright/', ws3_2)

write_squid('source/squid#1/stage00/head/bright/', ws3_2)
write_squid('source/squid#1/stage01/head/bright/', ws3_2)
write_squid('source/squid#1/stage02/head/bright/', ws3_2)

wb.save("./excel/test.xlsx")
"""
